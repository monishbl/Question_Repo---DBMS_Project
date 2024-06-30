from flask import Flask, request, redirect, url_for, render_template

import random
import openpyxl
from connect import get_db_connection

app = Flask(__name__)

qu_count = 0
count = 0
allowed_urls = ["/"]
questions = []
score = 0
username = ""
ADMIN_USERNAME = "admin"
error_message = ""

wb = openpyxl.load_workbook("interview_questions_mcq.xlsx")
ws = wb.active

interview_questions_mcq_dict = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    question, options, answer_index, difficulty = row
    options_list = options.split("\n")
    interview_questions_mcq_dict[question] = {
        "options": options_list,
        "answer_index": int(answer_index),
        "difficulty": difficulty,
    }


# Database query execution
def query_db(query, args=(), one=False):
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(query, args)
    rv = cursor.fetchall()
    cursor.close()
    conn.commit()
    conn.close()
    return (rv[0] if rv else None) if one else rv

# Fetch total number of questions
def get_total_questions():
    result = query_db("SELECT COUNT(*) as total FROM qa", one=True)
    return result["total"]

def select_random_question():
    total_questions = get_total_questions()
    all_questions = list(range(1, total_questions + 1))
    available_questions = [
        q for q in all_questions if not any(ques["qno"] == str(q) for ques in questions)
    ]
    if not available_questions:
        return random.choice(all_questions)
    ran = random.choice(available_questions)
    return ran

def check_answer(question, answer):
    if answer is None:
        return False

    try:
        answer_int = int(answer)
    except ValueError:
        return False

    data = query_db("SELECT answer FROM qa WHERE qno=%s;", (question,), one=True)
    return data and data["answer"] == answer_int

@app.route("/", methods=["GET", "POST"])
def home():
    global questions, score, error_message
    if request.method == "POST":
        login_type = request.form.get("login_type")
        if login_type == "user":
            return redirect(url_for("user_login"))
        elif login_type == "admin":
            return redirect(url_for("admin_login_form"))
        else:
            return "Invalid login choice"
    else:
        questions = []
        score = 0
        return render_template("index.html", error_message=error_message)

@app.route("/user_login", methods=["POST", "GET"])
def user_login():
    global username, allowed_urls
    allowed_urls.append("/user_login")
    if request.method == "POST":
        username = request.form.get("username")
        allowed_urls.append("/quiz")
        return redirect(url_for("quiz"))
    return render_template("index.html")

@app.route("/admin_authenticate", methods=["POST"])
def admin_authenticate():
    allowed_urls.append("/admin_authenticate")
    admin_username = request.form.get("admin_username")
    if admin_username != ADMIN_USERNAME:
        return "Access Denied", 403
    allowed_urls.append("/admin")
    return redirect(url_for("admin"))

@app.route("/quiz", methods=["GET", "POST"])
def quiz():
    global error_message, count, qu_count
    if "/quiz" not in allowed_urls:
        error_message = "Access Denied - Please login to continue."
        return redirect(url_for("home"))
    else:
        count += 1
        qu_count += 1
        global score, questions, username
        question = request.form.get("qno")
        user_answer = request.form.get("answer")
        count = int(request.form.get("count", 0)) + 1
        if (
            question
            and user_answer
            and not any(q["qno"] == question for q in questions)
        ):
            correct = check_answer(question, user_answer)
            if correct:
                score += 1
            questions.append(
                {
                    "qno": question,
                    "user_answer": user_answer,
                    "correct": correct,
                    "answer": query_db(
                        "SELECT choice%s FROM qa WHERE qno=%s;",
                        (
                            int(user_answer) + 1,
                            question,
                        ),
                        one=True,
                    )["choice%s" % (int(user_answer) + 1)],
                    "question": query_db(
                        "SELECT question FROM qa WHERE qno=%s;", (question,), one=True
                    )["question"],
                    "correct_choice": query_db(
                        "SELECT choice%s FROM qa WHERE qno=%s;",
                        (
                            int(
                                query_db(
                                    "SELECT answer FROM qa WHERE qno=%s;",
                                    (question,),
                                    one=True,
                                )["answer"]
                            )
                            + 1,
                            question,
                        ),
                        one=True,
                    )["choice%s" % (int(
                                query_db(
                                    "SELECT answer FROM qa WHERE qno=%s;",
                                    (question,),
                                    one=True,
                                )["answer"]
                            )
                            + 1)],
                }
            )
        total_questions = get_total_questions()

        if len(questions) >= total_questions or count-1 >= 10:
            allowed_urls.append("/result")
            return redirect(url_for("result"))
        ran = select_random_question()
        data = query_db("SELECT * FROM qa WHERE qno=%s;", (ran,), one=True)
        return render_template(
            "quiz.html",
            qu_count=qu_count,
            data=data,
            count=count,
            total_questions=total_questions,
            username=username,
        )

@app.route("/result", methods=["GET", "POST"])
def result():
    global error_message
    if "/result" not in allowed_urls:
        error_message = "Please login to continue and take the quiz."
        return redirect(url_for("home"))
    else:
        global score, questions, username
        return render_template(
            "result.html",
            score=score,
            questions=questions,
            username=username,
        )

@app.route("/admin", methods=["GET", "POST"])
def admin():
    global error_message
    if "/admin" not in allowed_urls:
        error_message = "Access Denied - Please login to continue."
        return redirect(url_for("home"))
    else:
        if request.method == "POST":
            action = request.form.get("action")
            qno = request.form.get("qno")
            question = request.form.get("question")
            answer = request.form.get("answer")
            choice1 = request.form.get("choice1")
            choice2 = request.form.get("choice2")
            choice3 = request.form.get("choice3")
            choice4 = request.form.get("choice4")


            if action == "add":
                query_db(
                    "INSERT INTO qa (qno, question, answer, choice1, choice2, choice3, choice4) VALUES (%s, %s, %s, %s, %s, %s, %s);",
                    (qno, question, answer, choice1, choice2, choice3, choice4),
                )
                wb = openpyxl.load_workbook("interview_questions_mcq.xlsx")
                ws = wb.active
                ws.append([question, "\n".join([choice1, choice2, choice3, choice4]), answer, "easy"])
                wb.save("interview_questions_mcq.xlsx")
                return "Question Added Successfully"
            elif action == "remove":
                query_db("DELETE FROM qa WHERE qno=%s;", (qno,))
                return "Question Removed Successfully"
            elif action == "modify":
                query_db(
                    "UPDATE qa SET question=%s, answer=%s, choice1=%s, choice2=%s, choice3=%s, choice4=%s WHERE qno=%s;",
                    (question, answer, choice1, choice2, choice3, choice4, qno),
                )
                #update in excel sheet
                wb = openpyxl.load_workbook("interview_questions_mcq.xlsx")
                ws = wb.active
                # Before the loop, find the index of the row you want to delete
                row_to_delete = None
                current_row_index = 2  # Assuming row 1 is headers and data starts from row 2
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] == question:
                        row_to_delete = current_row_index
                        break
                    current_row_index += 1

                # If a row to delete is found, delete it
                if row_to_delete is not None:
                    ws.delete_rows(row_to_delete)

                ws.append([question, "\n".join([choice1, choice2, choice3, choice4]), answer, "easy"])
                wb.save("interview_questions_mcq.xlsx")

                return "Question Modified Successfully"
            else:
                return "Invalid Action", 400

        return render_template("admin.html")

@app.route("/admin/question/<qno>", methods=["GET"])
def get_question(qno):
    question = query_db("SELECT * FROM qa WHERE qno=%s;", (qno,), one=True)
    if question:
        return question
    return {}, 404

if __name__ == "__main__":
    mydb = get_db_connection()
    cursor = mydb.cursor()
    cursor.execute("DROP TABLE IF EXISTS qa")
    cursor.execute("CREATE TABLE qa (qno INT PRIMARY KEY AUTO_INCREMENT, question VARCHAR(255), answer INT, difficulty varchar(255), choice1 VARCHAR(255), choice2 VARCHAR(255), choice3 VARCHAR(255), choice4 VARCHAR(255))")
    insert_query = "INSERT INTO qa (question, answer, difficulty, choice1, choice2, choice3, choice4) VALUES (%s, %s, %s, %s, %s, %s, %s)"
    for question, details in interview_questions_mcq_dict.items():
        data_tuple = (
            question,
            details["answer_index"],
            details["difficulty"],
            details["options"][0],
            details["options"][1],
            details["options"][2],
            details["options"][3],
        )
        cursor.execute(insert_query, data_tuple)
    mydb.commit()
    app.run(debug=True)